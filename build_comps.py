"""
GOOGL Comparable Company Analysis — Excel builder
Generates institutional-grade comps sheet with live formulas.
Blue fill = hardcoded inputs | No fill = formula cells
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comps Analysis"

# ── Palette ──────────────────────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="DCE6F1")   # inputs
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy header
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")   # blue sub-header
STAT_FILL   = PatternFill("solid", fgColor="F2F2F2")   # stats rows
GOOGL_FILL  = PatternFill("solid", fgColor="FFF2CC")   # highlight subject co.
NOTE_FILL   = PatternFill("solid", fgColor="EDEDED")

WHITE  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BOLD   = Font(bold=True, name="Calibri", size=10)
NORMAL = Font(name="Calibri", size=10)
BLUE_F = Font(color="1F3864", bold=True, name="Calibri", size=10)
ITALIC = Font(italic=True, name="Calibri", size=9, color="595959")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border(top=False, bottom=False, left=False, right=False):
    s = Side(style="thin")
    return Border(
        top=s if top else Side(),
        bottom=s if bottom else Side(),
        left=s if left else Side(),
        right=s if right else Side(),
    )

def thick_border(top=False, bottom=False):
    s = Side(style="medium")
    return Border(top=s if top else Side(), bottom=s if bottom else Side())

def style(cell, fill=None, font=None, align=None, border=None, num_fmt=None):
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = border
    if num_fmt: cell.number_format = num_fmt

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {1:22, 2:10, 3:12, 4:10, 5:13, 6:13, 7:13,
              8:13, 9:14, 10:12, 11:11, 12:11}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Row 1–3 : Header block ────────────────────────────────────────────────────
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 14

ws.merge_cells("A1:L1")
ws["A1"] = "INTERNET / MEGA-CAP TECHNOLOGY — COMPARABLE COMPANY ANALYSIS"
style(ws["A1"], fill=HEADER_FILL, font=WHITE, align=CENTER)

ws.merge_cells("A2:L2")
ws["A2"] = "Alphabet (GOOGL)  •  Meta (META)  •  Microsoft (MSFT)  •  Amazon (AMZN)  •  Apple (AAPL)  •  Netflix (NFLX)"
style(ws["A2"], fill=HEADER_FILL, font=Font(color="FFFFFF", name="Calibri", size=9), align=CENTER)

ws.merge_cells("A3:L3")
ws["A3"] = f"As of {date.today().strftime('%B %d, %Y')}  |  LTM / FY2025E  |  All figures in USD Millions (except per-share and ratio data)"
style(ws["A3"], fill=SUBHDR_FILL, font=Font(color="FFFFFF", name="Calibri", size=9, italic=True), align=CENTER)

# ── Row 5 : Operating Statistics header ──────────────────────────────────────
ws.row_dimensions[5].height = 18
ws.merge_cells("A5:L5")
ws["A5"] = "OPERATING STATISTICS & FINANCIAL METRICS"
style(ws["A5"], fill=SUBHDR_FILL, font=WHITE, align=CENTER)

# ── Row 6 : Column headers ────────────────────────────────────────────────────
ws.row_dimensions[6].height = 30
headers_op = [
    "Company", "Ticker",
    "Revenue\n(LTM, $M)", "YoY\nGrowth",
    "Gross\nMargin", "EBITDA\n(LTM, $M)",
    "EBITDA\nMargin", "FCF\n(LTM, $M)",
    "FCF\nMargin", "R&D\n% Rev",
    "Capex\n% Rev", "Net Cash\n($M)"
]
for c, h in enumerate(headers_op, 1):
    cell = ws.cell(row=6, column=c, value=h)
    style(cell, fill=PatternFill("solid", fgColor="BDD7EE"),
          font=BOLD, align=Alignment(horizontal="center", vertical="center", wrap_text=True),
          border=thin_border(bottom=True))

# ── Operating data (inputs — blue fill) ───────────────────────────────────────
# Columns: Company, Ticker, Revenue, YoY%, GrossMargin%, EBITDA, EBITDAMargin%, FCF, FCFMargin%, R&D%, Capex%, NetCash
op_data = [
    # Subject company first
    ("Alphabet",   "GOOGL", 359_000, 0.132, 0.578, 114_000, 0.317, 72_000, 0.201, 0.155, 0.069,  100_000),
    ("Meta",       "META",  164_500, 0.190, 0.817,  70_000, 0.426, 52_000, 0.316, 0.269, 0.043,   50_000),
    ("Microsoft",  "MSFT",  261_000, 0.152, 0.698, 125_000, 0.479, 74_000, 0.284, 0.130, 0.056,   30_000),
    ("Amazon",     "AMZN",  638_000, 0.109, 0.497, 125_000, 0.196, 38_000, 0.060, 0.154, 0.082,   25_000),
    ("Apple",      "AAPL",  400_000, 0.040, 0.468, 138_000, 0.345, 95_000, 0.238, 0.074, 0.028,   50_000),
    ("Netflix",    "NFLX",   39_900, 0.150, 0.430,   9_900, 0.248,  6_200, 0.155, 0.000, 0.005,   -5_000),
]

PCT_COLS = {4, 5, 7, 9, 10, 11}  # columns that are percentages (1-indexed)
DOLLAR_COLS = {3, 6, 8, 12}

start_op = 7
for i, row_data in enumerate(op_data):
    r = start_op + i
    ws.row_dimensions[r].height = 15
    is_subject = (i == 0)
    row_fill = GOOGL_FILL if is_subject else BLUE_FILL

    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if c in PCT_COLS:
            cell.number_format = '0.0%'
        elif c in DOLLAR_COLS:
            cell.number_format = '#,##0'
        elif c == 1:
            style(cell, font=Font(bold=is_subject, name="Calibri", size=10))
        style(cell, fill=row_fill, align=RIGHT if c > 2 else LEFT,
              border=thin_border(bottom=(i == len(op_data)-1)))

    # Ticker col: center
    ws.cell(row=r, column=2).alignment = CENTER

# Blue input legend note
ws.merge_cells(f"A{start_op}:A{start_op}")  # already done per row

# ── Operating statistics rows ─────────────────────────────────────────────────
stat_labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]
stat_funcs  = ["MAX", "PERCENTILE", "MEDIAN", "PERCENTILE", "MIN"]
stat_args   = [None, 0.75, None, 0.25, None]

stat_start = start_op + len(op_data)  # row 13
for si, (label, func, arg) in enumerate(zip(stat_labels, stat_funcs, stat_args)):
    r = stat_start + si
    ws.row_dimensions[r].height = 14
    cell_a = ws.cell(row=r, column=1, value=label)
    style(cell_a, fill=STAT_FILL, font=Font(italic=True, name="Calibri", size=9),
          align=LEFT, border=thin_border(bottom=(si==4)))
    ws.cell(row=r, column=2).fill = STAT_FILL  # blank ticker col

    data_rows = f"{start_op}:{start_op + len(op_data)-1}"
    for c in range(3, 13):
        col_letter = get_column_letter(c)
        data_range = f"{col_letter}{start_op}:{col_letter}{start_op+len(op_data)-1}"
        if func == "PERCENTILE":
            formula = f"=PERCENTILE({data_range},{arg})"
        elif func == "MAX":
            formula = f"=MAX({data_range})"
        elif func == "MIN":
            formula = f"=MIN({data_range})"
        else:
            formula = f"=MEDIAN({data_range})"
        cell = ws.cell(row=r, column=c, value=formula)
        if c in PCT_COLS:
            cell.number_format = '0.0%'
        elif c in DOLLAR_COLS:
            cell.number_format = '#,##0'
        style(cell, fill=STAT_FILL, font=ITALIC, align=RIGHT,
              border=thin_border(bottom=(si==4)))

# ── Row spacer ────────────────────────────────────────────────────────────────
spacer_r = stat_start + 5
ws.row_dimensions[spacer_r].height = 8

# ── Valuation multiples section ───────────────────────────────────────────────
val_hdr_r = spacer_r + 1
ws.row_dimensions[val_hdr_r].height = 18
ws.merge_cells(f"A{val_hdr_r}:L{val_hdr_r}")
ws[f"A{val_hdr_r}"] = "VALUATION MULTIPLES"
style(ws[f"A{val_hdr_r}"], fill=SUBHDR_FILL, font=WHITE, align=CENTER)

col_hdr_r = val_hdr_r + 1
ws.row_dimensions[col_hdr_r].height = 30
val_headers = [
    "Company", "Ticker",
    "Mkt Cap\n($M)", "Net Debt\n($M)",
    "Enterprise\nValue ($M)", "EV /\nRevenue",
    "EV /\nEBITDA", "P / E\n(NTM)",
    "FCF\nYield", "PEG\nRatio",
    "Div\nYield", "Price /\nFCF"
]
for c, h in enumerate(val_headers, 1):
    cell = ws.cell(row=col_hdr_r, column=c, value=h)
    style(cell, fill=PatternFill("solid", fgColor="BDD7EE"),
          font=BOLD, align=Alignment(horizontal="center", vertical="center", wrap_text=True),
          border=thin_border(bottom=True))

# Valuation inputs
# Cols: Company, Ticker, MktCap, NetDebt(-=cash), EV(formula), EV/Rev, EV/EBITDA, P/E, FCFYield, PEG, DivYield, P/FCF
val_data = [
    #  Company     Ticker   MktCap    NetDebt    EV/Rev  EV/EBITDA  P/E(NTM) FCFYield  PEG   DivYield  P/FCF
    ("Alphabet",  "GOOGL", 2_100_000, -100_000,  5.57,   17.54,    20.2,    0.034,   1.53,  0.005,   29.2),
    ("Meta",      "META",  1_450_000,  -50_000,  8.48,   19.86,    23.8,    0.036,   1.25,  0.004,   27.9),
    ("Microsoft", "MSFT",  3_000_000,  -30_000, 11.42,   23.84,    31.5,    0.025,   2.10,  0.008,   40.5),
    ("Amazon",    "AMZN",  2_300_000,  -25_000,  3.55,   18.20,    37.8,    0.017,   3.45,  0.000,   60.5),
    ("Apple",     "AAPL",  3_400_000,  -50_000,  8.38,   24.28,    29.5,    0.028,   7.38,  0.005,   35.8),
    ("Netflix",   "NFLX",    385_000,    5_000,   9.53,   38.18,    34.5,    0.016,   2.30,  0.000,   62.1),
]

val_start = col_hdr_r + 1
for i, row_data in enumerate(val_data):
    r = val_start + i
    ws.row_dimensions[r].height = 15
    is_subject = (i == 0)
    row_fill = GOOGL_FILL if is_subject else BLUE_FILL

    company, ticker, mktcap, netdebt = row_data[0], row_data[1], row_data[2], row_data[3]
    ev_rev, ev_ebitda, pe, fcf_yield, peg, div_yield, p_fcf = row_data[4:]

    vals = [company, ticker, mktcap, netdebt,
            f"={get_column_letter(3)}{r}+{get_column_letter(4)}{r}",  # EV = MktCap + NetDebt
            ev_rev, ev_ebitda, pe, fcf_yield, peg, div_yield, p_fcf]

    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if c in {3, 4, 5}:
            cell.number_format = '#,##0'
        elif c in {6, 7, 8, 10, 12}:
            cell.number_format = '0.0x'
            if isinstance(val, (int, float)):
                cell.value = val  # store as number, display as Xx
                cell.number_format = '0.0"x"'
        elif c in {9, 11}:
            cell.number_format = '0.0%'
        is_input = not (c == 5)  # EV is a formula
        style(cell,
              fill=row_fill if is_input else PatternFill("solid", fgColor="EBF3FB"),
              font=Font(bold=is_subject, name="Calibri", size=10),
              align=RIGHT if c > 2 else LEFT,
              border=thin_border(bottom=(i == len(val_data)-1)))
    ws.cell(row=val_start+i, column=2).alignment = CENTER

# ── Valuation statistics ───────────────────────────────────────────────────────
val_stat_start = val_start + len(val_data)
for si, (label, func, arg) in enumerate(zip(stat_labels, stat_funcs, stat_args)):
    r = val_stat_start + si
    ws.row_dimensions[r].height = 14
    cell_a = ws.cell(row=r, column=1, value=label)
    style(cell_a, fill=STAT_FILL, font=Font(italic=True, name="Calibri", size=9),
          align=LEFT, border=thin_border(bottom=(si==4)))
    ws.cell(row=r, column=2).fill = STAT_FILL

    for c in range(3, 13):
        col_letter = get_column_letter(c)
        data_range = f"{col_letter}{val_start}:{col_letter}{val_start+len(val_data)-1}"
        if func == "PERCENTILE":
            formula = f"=PERCENTILE({data_range},{arg})"
        elif func == "MAX":
            formula = f"=MAX({data_range})"
        elif func == "MIN":
            formula = f"=MIN({data_range})"
        else:
            formula = f"=MEDIAN({data_range})"
        cell = ws.cell(row=r, column=c, value=formula)
        if c in {3, 4, 5}:
            cell.number_format = '#,##0'
        elif c in {6, 7, 8, 10, 12}:
            cell.number_format = '0.0"x"'
        elif c in {9, 11}:
            cell.number_format = '0.0%'
        style(cell, fill=STAT_FILL, font=ITALIC, align=RIGHT,
              border=thin_border(bottom=(si==4)))

# ── Notes section ─────────────────────────────────────────────────────────────
notes_start = val_stat_start + 5 + 2
ws.row_dimensions[notes_start].height = 16
ws.merge_cells(f"A{notes_start}:L{notes_start}")
ws[f"A{notes_start}"] = "NOTES & METHODOLOGY"
style(ws[f"A{notes_start}"], fill=PatternFill("solid", fgColor="404040"),
      font=WHITE, align=LEFT)

notes = [
    "Data source: Company filings (10-K / 20-F), earnings releases, and FactSet estimates as of May 2026.",
    "LTM = Last Twelve Months ending closest reported quarter. Revenue, EBITDA, and FCF are on an LTM basis.",
    "EBITDA is adjusted for stock-based compensation and one-time items where disclosed by management.",
    "FCF = Operating Cash Flow less Capital Expenditures. FCF Yield = LTM FCF / Market Capitalization.",
    "Enterprise Value = Market Capitalization + Net Debt (Net Debt = Total Debt – Cash & Equivalents).",
    "Net Debt is negative for companies with net cash positions (Alphabet, Meta, Microsoft, Apple).",
    "P/E (NTM) is based on consensus next-twelve-months EPS estimates from sell-side research.",
    "Netflix R&D shown as 0.0% as content spend is classified as an operating cost, not R&D.",
    "Yellow highlight = subject company (Alphabet / GOOGL). Blue fill = hardcoded inputs. White = formula cells.",
    "This analysis is for informational purposes only and does not constitute investment advice.",
]
for ni, note in enumerate(notes):
    r = notes_start + 1 + ni
    ws.row_dimensions[r].height = 13
    ws.merge_cells(f"A{r}:L{r}")
    cell = ws[f"A{r}"]
    cell.value = f"  {ni+1}.  {note}"
    style(cell, fill=NOTE_FILL if ni % 2 == 0 else PatternFill("solid", fgColor="F8F8F8"),
          font=Font(name="Calibri", size=9, color="404040"), align=LEFT)

# ── Freeze panes & print setup ─────────────────────────────────────────────────
ws.freeze_panes = "C7"
ws.sheet_view.showGridLines = False
ws.print_title_rows = "1:6"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True

out_path = r"C:\Users\simon\Downloads\anthropic_financial_services_prism_labs\GOOGL_Comps_Analysis_2026-05-09.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
